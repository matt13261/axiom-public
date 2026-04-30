# AXIOM — train.py
# Script d'entraînement complet : MCCFR blueprint + Deep CFR.
# Mise à jour Phase 10 : évaluation semi-pro (TAG, LAG, Régulier).
# =============================================================================
#
# Usage :
#     python train.py --mode mccfr --iterations 100000   # test rapide
#     python train.py --mode mccfr --iterations 1000000  # entraînement complet
#     python train.py --mode deepcfr
#     python train.py --mode eval                         # évaluation complète
#     python train.py --mode eval --nb-mains-eval 5000   # plus précis
#     python train.py --mode bench
#     python train.py --workers 8                        # forcer nb de workers
#
# Le MCCFR utilise automatiquement tous les cœurs CPU disponibles.
# Sur Ryzen 7 5800HS (8 cœurs physiques) : gain ~6-8x.
#
# Fichiers générés :
#     data/strategy/blueprint_v1.pkl
#     data/models/strategy_net_j*.pt
#     data/logs/training_log.csv
#     data/logs/progression_mccfr.csv
#
# =============================================================================

import os
import sys
import time
import argparse
import csv
import shutil
import threading
import multiprocessing as mp
from datetime import datetime

try:
    from tqdm import tqdm as _tqdm
    _TQDM_OK = True
except ImportError:
    _TQDM_OK = False


# ---------------------------------------------------------------------------
# Répartition des itérations MCCFR par niveau de blinde
#
# N°6 + N°13 : redistribution en courbe en cloche centrée sur 20/40–40/80.
#
# Ancienne répartition (pyramide décroissante) :
#   10/20=35%, 15/30=25%, 20/40=15%, 30/60=10%, 40/80=6%, 50/100=4%,
#   60/120=3%, 80/160=2%
#   → Surcharge les premiers niveaux (stacks profonds, peu stratégiques)
#   → Seulement 2% pour 80/160 alors que c'est la situation la plus
#     fréquente en fin de tournoi (ratio 3 BB — push/fold critique)
#
# Nouvelle répartition (courbe en cloche) :
#   → Pic sur 20/40 et 30/60 (ratio 12–16 BB : zone stratégiquement riche,
#     open/3-bet/squeeze encore possibles)
#   → Plus d'itérations sur 40/80–80/160 (stacks courts, phases décisives)
#   → Moins sur 10/20 (ratio 25 BB initial, trop profond pour le tournoi)
#
# Ratio stack/BB pour STACK_DEPART=500 :
#   10/20 = 25 BB | 15/30 = 16 BB | 20/40 = 12 BB | 30/60 = 8 BB
#   40/80 = 6 BB  | 50/100 = 5 BB | 60/120 = 4 BB | 80/160 = 3 BB
# ---------------------------------------------------------------------------

_NIVEAUX_ENTRAINEMENT = [
    ( 10,  20, 0.22),   # 25 BB  — complexité maximale (postflop complet, 4 streets)
    ( 15,  30, 0.25),   # 16 BB  — pic stratégique (3-bet/squeeze encore viables)
    ( 20,  40, 0.20),   # 12 BB  — transition push/fold (open-shove commence)
    ( 30,  60, 0.15),   # 8 BB   — push/fold dominant
    ( 40,  80, 0.08),   # 6 BB   — quasi résolu rapidement
    ( 50, 100, 0.05),   # 5 BB   — push/fold pur
    ( 60, 120, 0.03),   # 4 BB   — trivial
    ( 80, 160, 0.02),   # 3 BB   — trivial (shove all-in systématique)
]


# ---------------------------------------------------------------------------
# Création des dossiers
# ---------------------------------------------------------------------------

def creer_dossiers() -> None:
    for d in ["data", "data/strategy", "data/models", "data/logs"]:
        os.makedirs(d, exist_ok=True)


# ---------------------------------------------------------------------------
# Affichage
# ---------------------------------------------------------------------------

def _titre(t)  : print("\n" + "="*58 + f"\n  {t}\n" + "="*58)
def _section(t): print(f"\n{'─'*58}\n  {t}\n{'─'*58}")
def _ok(t)     : print(f"  OK  {t}")
def _info(t)   : print(f"  >>  {t}")
def _warn(t)   : print(f"  !!  {t}")


# ---------------------------------------------------------------------------
# CSV progression
# ---------------------------------------------------------------------------

def _ecrire_rapport(chemin_csv: str, ligne: dict) -> None:
    ecrire_entete = not os.path.exists(chemin_csv)
    with open(chemin_csv, 'a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=list(ligne.keys()))
        if ecrire_entete:
            writer.writeheader()
        writer.writerow(ligne)


# ---------------------------------------------------------------------------
# WORKER — spawné UNE SEULE FOIS par niveau, fait toutes ses itérations
# ---------------------------------------------------------------------------

_TRANCHE_WORKER = 500   # itérations par tranche → granularité du compteur


def _worker_mccfr(pb: int, gb: int, nb_iterations: int,
                  stack: int, queue: mp.Queue, worker_id: int,
                  compteur=None, biais_conf: tuple = None) -> None:
    """
    Exécuté dans un processus séparé.
    Reçoit TOUTES ses itérations d'un coup — pas de re-spawn en cours de route.
    Retourne son dictionnaire de noeuds à la fin.

    biais_conf : tuple (categorie, alpha) ou None → baseline.
                 Ex: ('raise', 0.05) pour la variante agressive Pluribus.

    compteur : mp.Value('i') partagé avec le processus principal pour la
               barre de progression. None = pas de suivi.
    """
    dossier = os.path.dirname(os.path.abspath(__file__))
    if dossier not in sys.path:
        sys.path.insert(0, dossier)

    from ai.mccfr import MCCFRHoldEm, BiaisContinuation

    # Point 2 — Continuation Strategies : instancier le biais dans le worker
    # (les objets BiaisContinuation eux-mêmes passeraient le pickle, mais on
    # reste conservateur en ne transmettant qu'un tuple simple).
    biais = None
    if biais_conf is not None:
        cat, alpha = biais_conf
        if cat is not None:
            biais = BiaisContinuation(cat, alpha=alpha)

    mccfr = MCCFRHoldEm(biais=biais)
    if nb_iterations > 0:
        if compteur is not None:
            # Traiter par tranches pour mettre à jour le compteur partagé
            restant = nb_iterations
            while restant > 0:
                tranche = min(_TRANCHE_WORKER, restant)
                mccfr.entrainer(
                    nb_iterations=tranche,
                    stacks=stack, pb=pb, gb=gb,
                    verbose=False, save_every=0,
                )
                with compteur.get_lock():
                    compteur.value += tranche
                restant -= tranche
        else:
            mccfr.entrainer(
                nb_iterations=nb_iterations,
                stacks=stack, pb=pb, gb=gb,
                verbose=False, save_every=0,
            )

    queue.put((worker_id, mccfr.noeuds))


# ---------------------------------------------------------------------------
# FUSION des dictionnaires de noeuds
# ---------------------------------------------------------------------------

def _fusionner_noeuds(liste_noeuds: list) -> dict:
    """
    Fusionne N dictionnaires NoeudCFR en additionnant regrets et stratégies.
    Valide mathématiquement : les regrets CFR sont additifs.
    """
    from ai.mccfr import NoeudCFR

    fusionne = {}
    for noeuds in liste_noeuds:
        for cle, noeud in noeuds.items():
            if cle not in fusionne:
                nouveau = NoeudCFR(noeud.nb_actions)
                nouveau.regrets_cumules = list(noeud.regrets_cumules)
                nouveau.strategie_somme = list(noeud.strategie_somme)
                nouveau.nb_visites      = noeud.nb_visites
                fusionne[cle]           = nouveau
            else:
                n = fusionne[cle]
                if noeud.nb_actions > n.nb_actions:
                    diff = noeud.nb_actions - n.nb_actions
                    n.regrets_cumules.extend([0.0] * diff)
                    n.strategie_somme.extend([0.0] * diff)
                    n.nb_actions = noeud.nb_actions
                for i in range(min(noeud.nb_actions, n.nb_actions)):
                    n.regrets_cumules[i] += noeud.regrets_cumules[i]
                    n.strategie_somme[i] += noeud.strategie_somme[i]
                n.nb_visites += noeud.nb_visites

    return fusionne


# ---------------------------------------------------------------------------
# PHASE 1 — Blueprint MCCFR (multiprocessing corrigé)
# ---------------------------------------------------------------------------

def entrainer_mccfr(iterations_total: int, sauvegarder_tous: int,
                    nb_workers: int = None,
                    biais_conf: tuple = None) -> dict:
    """
    Lance l'entraînement MCCFR en parallèle sur tous les cœurs CPU.
    Workers spawnés UNE SEULE FOIS par niveau — pas de re-spawn intermédiaire.

    biais_conf : tuple (categorie, alpha) ou None. Produit une variante de
                 blueprint pour les Continuation Strategies (Pluribus k=4).
                 Le suffixe "_<categorie>" est ajouté au chemin de sortie.
    """
    _section("MCCFR — Génération du Blueprint (multiprocessing)")

    from ai.strategy import sauvegarder_blueprint
    from config.settings import CHEMIN_BLUEPRINT, STACK_DEPART, CHEMIN_LOG, NB_WORKERS_MAX

    # Chemin de sortie : ajouter le suffixe biais si non-baseline
    chemin_sortie = CHEMIN_BLUEPRINT
    suffixe_biais = ""
    if biais_conf is not None and biais_conf[0] is not None:
        cat = biais_conf[0]
        suffixe_biais = f"_{cat}"
        base, ext = os.path.splitext(CHEMIN_BLUEPRINT)
        chemin_sortie = f"{base}{suffixe_biais}{ext}"

    chemin_progression = CHEMIN_LOG.replace("training_log.csv",
                                            "progression_mccfr.csv")

    cpu_dispo  = mp.cpu_count()
    if nb_workers is None:
        nb_workers = min(max(1, cpu_dispo - 1), NB_WORKERS_MAX)
    nb_workers = min(nb_workers, cpu_dispo, NB_WORKERS_MAX)

    plan = []
    total_alloue = 0
    for i, (pb, gb, pct) in enumerate(_NIVEAUX_ENTRAINEMENT):
        nb = int(iterations_total * pct)
        if i == len(_NIVEAUX_ENTRAINEMENT) - 1:
            nb = iterations_total - total_alloue
        total_alloue += nb
        plan.append((pb, gb, nb))

    _info(f"Stack de départ    : {STACK_DEPART} jetons")
    _info(f"Iterations totales : {iterations_total:,}")
    _info(f"Workers CPU        : {nb_workers} / {cpu_dispo} cœurs")
    if biais_conf is not None and biais_conf[0] is not None:
        _info(f"Biais continuation : {biais_conf[0]} (alpha={biais_conf[1]})")
    _info(f"Destination        : {chemin_sortie}")
    print()
    print(f"  {'Blindes':>8} | {'BB':>5} | {'Total':>12} | {'Par worker':>12}")
    print(f"  {'─'*8}-+-{'─'*5}-+-{'─'*12}-+-{'─'*12}")
    for pb, gb, nb in plan:
        print(f"  {pb:>3}/{gb:<4} | {STACK_DEPART//gb:>5} | "
              f"{nb:>12,} | {nb//nb_workers:>12,}")
    print()

    it_par_sec_estime = 50 * nb_workers
    duree_estimee_min = iterations_total / it_par_sec_estime / 60
    print(f"  Durée estimée : ~{duree_estimee_min:.0f} min "
          f"({duree_estimee_min/60:.1f}h)\n")

    noeuds_total        = {}
    debut_global        = time.time()
    it_global           = 0
    derniere_sauvegarde = 0

    for niveau_idx, (pb, gb, nb_iterations) in enumerate(plan):

        bb_ratio     = STACK_DEPART / gb
        debut_niveau = time.time()

        print(f"  Niveau {niveau_idx+1}/8 : blindes {pb}/{gb} "
              f"({bb_ratio:.0f} BB) — {nb_iterations:,} itérations")

        # Répartir les itérations entre workers
        its_par_worker = [nb_iterations // nb_workers] * nb_workers
        reste = nb_iterations % nb_workers
        for i in range(reste):
            its_par_worker[i] += 1

        # Compteur partagé pour la barre de progression
        compteur = mp.Value('i', 0)

        result_queue = mp.Queue()
        workers      = []
        for w_id, nb_it_w in enumerate(its_par_worker):
            p = mp.Process(
                target=_worker_mccfr,
                args=(pb, gb, nb_it_w, STACK_DEPART, result_queue, w_id,
                      compteur, biais_conf),
                daemon=True,
            )
            p.start()
            workers.append(p)

        # ── Barre de progression (thread daemon) ───────────────────────
        stop_prog = threading.Event()
        if _TQDM_OK:
            desc = f"  Niv.{niveau_idx+1}/8 {pb:>2}/{gb}"
            def _thread_prog(stop, cpt, total, desc=desc):
                fmt = ('{desc}: {percentage:3.0f}%|{bar}| '
                       '{n_fmt}/{total_fmt} it '
                       '[{elapsed}<{remaining}, {rate_fmt}]')
                with _tqdm(total=total, desc=desc, unit='it',
                           unit_scale=True, ncols=78,
                           bar_format=fmt, leave=True) as pbar:
                    dernier = 0
                    while not stop.is_set():
                        courant = cpt.value
                        if courant > dernier:
                            pbar.update(courant - dernier)
                            dernier = courant
                        time.sleep(0.35)
                    # Flush final
                    courant = cpt.value
                    if courant > dernier:
                        pbar.update(courant - dernier)
            t_prog = threading.Thread(
                target=_thread_prog, args=(stop_prog, compteur, nb_iterations),
                daemon=True,
            )
            t_prog.start()

        # Collecter les résultats
        resultats = {}
        for _ in workers:
            w_id, noeuds = result_queue.get()
            resultats[w_id] = noeuds

        for p in workers:
            p.join()

        # Arrêter la barre de progression
        stop_prog.set()
        if _TQDM_OK:
            t_prog.join()

        # Fusionner
        liste_noeuds = [resultats[i] for i in range(nb_workers)]
        noeuds_niveau = _fusionner_noeuds(liste_noeuds)

        # Fusionner dans le total
        noeuds_total = _fusionner_noeuds([noeuds_total, noeuds_niveau])

        it_global  += nb_iterations
        nb_noeuds   = len(noeuds_total)
        duree_niveau = time.time() - debut_niveau
        it_par_sec   = nb_iterations / max(duree_niveau, 0.001)

        print(f"    ✓ {nb_iterations:,} it. | {it_par_sec:.0f} it/s | "
              f"{nb_noeuds:,} infosets | {duree_niveau:.1f}s")

        # Rapport CSV
        _ecrire_rapport(chemin_progression, {
            "timestamp"   : datetime.now().isoformat(timespec='seconds'),
            "niveau"      : f"{pb}/{gb}",
            "bb"          : bb_ratio,
            "iterations"  : it_global,
            "infosets"    : nb_noeuds,
            "it_par_sec"  : round(it_par_sec),
            "duree_s"     : round(duree_niveau, 1),
        })

        # Sauvegarder le blueprint après chaque niveau
        sauvegarder_blueprint(noeuds_total, chemin_sortie)
        derniere_sauvegarde = it_global
        print(f"  >>> Blueprint sauvegardé ({it_global:,} it. | "
              f"{nb_noeuds:,} infosets)")

    print()
    duree_totale = time.time() - debut_global
    _ok(f"Blueprint sauvegardé   : {chemin_sortie}")
    _ok(f"Infosets générés       : {len(noeuds_total):,}")
    _ok(f"Workers utilisés       : {nb_workers}")
    _ok(f"Rapport de progression : {chemin_progression}")
    _ok(f"Durée totale           : {duree_totale:.1f}s  ({duree_totale/60:.1f} min)")

    return noeuds_total


# ---------------------------------------------------------------------------
# PHASE 2 — Deep CFR (parallèle — 3 processus simultanés, un par joueur)
# ---------------------------------------------------------------------------

# Taille du reservoir buffer par joueur
_TAILLE_BUFFER_DEEPCFR = 3_000_000

# Dossier des checkpoints numérotés (relatif au script)
_DOSSIER_CHECKPOINTS = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "checkpoints"
)
_DOSSIER_BUFFERS = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "buffers"
)
_CHEMIN_EXCEL_EVAL = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "logs", "eval_iterations.xlsx"
)


def _trouver_dernier_checkpoint() -> tuple:
    """
    Scanne data/checkpoints/ et retourne (iteration, chemin_dossier)
    du checkpoint le plus récent.
    Retourne (0, None) si aucun checkpoint n'existe.
    """
    if not os.path.exists(_DOSSIER_CHECKPOINTS):
        return 0, None

    dossiers = [
        d for d in os.listdir(_DOSSIER_CHECKPOINTS)
        if os.path.isdir(os.path.join(_DOSSIER_CHECKPOINTS, d))
        and d.startswith("iteration_")
    ]
    if not dossiers:
        return 0, None

    candidats = []
    for d in dossiers:
        try:
            num = int(d.split("_")[1])
            candidats.append((num, os.path.join(_DOSSIER_CHECKPOINTS, d)))
        except (IndexError, ValueError):
            continue

    if not candidats:
        return 0, None

    return max(candidats, key=lambda x: x[0])


def _worker_traversees(joueur_i: int, nb_trav: int, stacks: int,
                        iteration: int, queue: mp.Queue) -> None:
    """
    Worker multiprocessing : fait nb_trav traversées pour joueur_i.
    Retourne les échantillons collectés (regrets + stratégie) via la queue.
    Les 3 workers tournent en parallèle — un par joueur.
    """
    dossier = os.path.dirname(os.path.abspath(__file__))
    if dossier not in sys.path:
        sys.path.insert(0, dossier)

    from ai.deep_cfr import DeepCFR

    dcfr = DeepCFR(taille_buffer=1)
    dcfr.iteration = iteration
    try:
        dcfr.charger(verbose=False)
    except Exception:
        pass

    echant_regret    = []
    echant_strategie = []

    class _BufferCollecteur:
        def __init__(self, liste):
            self._liste = liste
        def ajouter(self, vec, data, *extra):
            self._liste.append((vec.copy(), data.copy()) + extra)
        def __len__(self):
            return len(self._liste)

    for j in range(3):
        dcfr.buffers_regret[j]    = _BufferCollecteur(echant_regret    if j == joueur_i else [])
        dcfr.buffers_strategie[j] = _BufferCollecteur(echant_strategie if j == joueur_i else [])

    for _ in range(nb_trav):
        etat = dcfr._dealer_aleatoire(stacks, 10, 20)
        dcfr._traverser(etat, joueur_i, iteration)

    queue.put((joueur_i, echant_regret, echant_strategie))


def entrainer_deep_cfr(iterations: int, traversees: int) -> None:
    """
    Deep CFR avec traversées parallèles (3 processus simultanés),
    buffers persistants et checkpoints numérotés par itération.
    """
    _section("Deep CFR — Entraînement Parallèle (3 processus)")

    import torch
    from ai.deep_cfr import DeepCFR
    from ai.trainer import NB_BATCHS_PAR_EPOCH
    from config.settings import (
        CHEMIN_REGRET_NET, CHEMIN_STRATEGY_NET,
        BATCH_SIZE, LEARNING_RATE, STACK_DEPART,
    )
    from ai.network import DEVICE

    _info(f"Iterations     : {iterations}")
    _info(f"Traversées/it  : {traversees} (par joueur, en parallèle)")
    _info(f"Buffer/joueur  : {_TAILLE_BUFFER_DEEPCFR:,} échantillons")
    _info(f"Batch size     : {BATCH_SIZE}")
    _info(f"Learning rate  : {LEARNING_RATE}")
    _info(f"Device         : {DEVICE}")
    _info(f"Stack          : {STACK_DEPART} jetons")
    print()

    os.makedirs(_DOSSIER_CHECKPOINTS, exist_ok=True)
    os.makedirs(_DOSSIER_BUFFERS,     exist_ok=True)

    # Détecter le dernier checkpoint et y reprendre l'itération
    num_checkpoint, chemin_checkpoint = _trouver_dernier_checkpoint()
    if chemin_checkpoint:
        _info(f"Checkpoint détecté : iteration_{num_checkpoint:03d} — copie vers data/models/")
        for f in os.listdir(chemin_checkpoint):
            if f.endswith(".pt"):
                shutil.copy2(os.path.join(chemin_checkpoint, f),
                             os.path.join("data", "models", f))
    else:
        _info("Aucun checkpoint — départ à l'itération 0")

    dcfr = DeepCFR(taille_buffer=_TAILLE_BUFFER_DEEPCFR)

    # Charger les réseaux (depuis checkpoint copié ou data/models existants)
    try:
        dcfr.charger(verbose=False)
        _info("Réseaux .pt chargés depuis la session précédente")
    except Exception:
        _info("Départ à zéro (pas de .pt existants)")

    # Partir du bon numéro d'itération
    dcfr.iteration = num_checkpoint
    if num_checkpoint > 0:
        _info(f"Reprise depuis l'itération {num_checkpoint}")

    # Charger les buffers persistants
    _info("Chargement des buffers...")
    for j in range(3):
        ok_r = dcfr.buffers_regret[j].charger(
            os.path.join(_DOSSIER_BUFFERS, f"buffer_regret_j{j}"))
        ok_s = dcfr.buffers_strategie[j].charger(
            os.path.join(_DOSSIER_BUFFERS, f"buffer_strategie_j{j}"))
        statut = "repris" if (ok_r and ok_s) else "vide"
        print(f"    J{j} : regrets={len(dcfr.buffers_regret[j]):,}  "
              f"stratégie={len(dcfr.buffers_strategie[j]):,}  [{statut}]")
    print()

    temps_trav_total  = 0.0
    temps_train_total = 0.0
    t_global          = time.time()

    for it in range(1, iterations + 1):
        dcfr.iteration += 1
        t_it = time.time()

        # ── Traversées en parallèle (3 processus) ──────────────────────────
        t0    = time.time()
        queue = mp.Queue()
        procs = [
            mp.Process(
                target=_worker_traversees,
                args=(j, traversees, STACK_DEPART, dcfr.iteration, queue),
                daemon=True,
            )
            for j in range(3)
        ]
        for p in procs:
            p.start()

        resultats = {}
        for _ in range(3):
            joueur_i, er, es = queue.get()
            resultats[joueur_i] = (er, es)
        for p in procs:
            p.join()

        t_trav = time.time() - t0
        temps_trav_total += t_trav

        # Insérer dans les buffers
        for j in range(3):
            er, es = resultats[j]
            nb_nouveaux = len(er)
            total_avant = dcfr.buffers_regret[j].nb_total
            for (vec, regrets, nb_actions) in er:
                dcfr.buffers_regret[j].ajouter(vec, regrets, nb_actions)
            for (vec, strat, iter_num, nb_actions) in es:
                dcfr.buffers_strategie[j].ajouter(vec, strat, iter_num, nb_actions)
            slots = len(dcfr.buffers_regret[j])
            total_apres = dcfr.buffers_regret[j].nb_total
            print(f"    J{j} | +{nb_nouveaux:6,} générés → slots={slots:,} | total_vu={total_apres:,}")

        print(f"  Traversées : {t_trav:.1f}s")

        # ── Entraînement des réseaux ────────────────────────────────────────
        # Réinitialiser le LR à chaque itération Deep CFR :
        # ReduceLROnPlateau est inadapté ici car les cibles (regrets) changent
        # à chaque itération → la loss fluctue toujours → le scheduler
        # pénalise à tort et tue le LR progressivement.
        # Point 10 — LR décroissant global : passe le numéro d'itération courante
        # pour que reinitialiser_scheduler() applique LR / sqrt(t)
        for j in range(3):
            dcfr.entraineurs_regret[j].reinitialiser_scheduler(
                iteration_courante=it, nb_iterations_total=iterations)
            dcfr.entraineurs_strategie[j].reinitialiser_scheduler(
                iteration_courante=it, nb_iterations_total=iterations)

        if torch.cuda.is_available():
            torch.cuda.synchronize()
        t0 = time.time()

        stats_r, stats_s = [], []
        for j in range(3):
            sr = dcfr.entraineurs_regret[j].entrainer_epoch(
                dcfr.buffers_regret[j],
                nb_batchs=NB_BATCHS_PAR_EPOCH, batch_size=BATCH_SIZE)
            ss = dcfr.entraineurs_strategie[j].entrainer_epoch(
                dcfr.buffers_strategie[j],
                nb_batchs=NB_BATCHS_PAR_EPOCH, batch_size=BATCH_SIZE)
            stats_r.append(sr)
            stats_s.append(ss)

        if torch.cuda.is_available():
            torch.cuda.synchronize()
        t_train = time.time() - t0
        temps_train_total += t_train

        # Tableau des pertes
        duree_it = time.time() - t_it
        print(f"\n  Itération Deep CFR #{dcfr.iteration}")
        print(f"  {'Joueur':>6} | {'Regrets (MSE)':>16} | {'Stratégie (MSE)':>15} | {'LR':>10}")
        print(f"  {'─'*6}-+-{'─'*16}-+-{'─'*15}-+-{'─'*10}")
        for j in range(3):
            lr    = dcfr.entraineurs_regret[j].optimiseur.param_groups[0]['lr']
            mse_r = stats_r[j].get('perte_moy', float('nan')) if isinstance(stats_r[j], dict) else float('nan')
            mse_s = stats_s[j].get('perte_moy', float('nan')) if isinstance(stats_s[j], dict) else float('nan')
            print(f"  {j:>6} | {mse_r:>16.6f} | {mse_s:>15.6f} | {lr:>10.2e}")

        t_total_elapsed = time.time() - t_global
        pct_trav  = 100 * t_trav  / max(duree_it, 0.001)
        pct_train = 100 * t_train / max(duree_it, 0.001)
        print(f"\n  Durée itération : {duree_it:.1f}s | Total : {t_total_elapsed:.0f}s")
        print(f"  Traversées : {t_trav:.1f}s ({pct_trav:.0f}%)  |  "
              f"Réseaux : {t_train:.1f}s ({pct_train:.0f}%)")

        # Sauvegarder les réseaux courants
        dcfr.sauvegarder(verbose=False)

        # Checkpoint numéroté
        dossier_it = os.path.join(_DOSSIER_CHECKPOINTS, f"iteration_{dcfr.iteration:03d}")
        os.makedirs(dossier_it, exist_ok=True)
        nb_copies = 0
        for f in os.listdir("data/models"):
            if f.endswith(".pt"):
                shutil.copy2(os.path.join("data/models", f),
                             os.path.join(dossier_it, f))
                nb_copies += 1
        print(f"  Checkpoint → iteration_{dcfr.iteration:03d}\\ ({nb_copies} .pt)")

        # Sauvegarder les buffers
        print(f"  Buffers...", end=" ", flush=True)
        t_buf = time.time()
        for j in range(3):
            dcfr.buffers_regret[j].sauvegarder(
                os.path.join(_DOSSIER_BUFFERS, f"buffer_regret_j{j}"))
            dcfr.buffers_strategie[j].sauvegarder(
                os.path.join(_DOSSIER_BUFFERS, f"buffer_strategie_j{j}"))
        taille_mo = sum(
            os.path.getsize(os.path.join(_DOSSIER_BUFFERS, f))
            for f in os.listdir(_DOSSIER_BUFFERS) if f.endswith(".npz")
        ) // (1024 * 1024)
        print(f"{time.time()-t_buf:.1f}s — {taille_mo} Mo")
        print()

    # Résumé
    t_total      = time.time() - t_global
    total_mesure = temps_trav_total + temps_train_total
    _ok(f"Réseaux sauvegardés : {CHEMIN_STRATEGY_NET.replace('.pt', '_j*.pt')}")
    _ok(f"Traversées  : {temps_trav_total:.1f}s ({100*temps_trav_total/max(total_mesure,1):.0f}%)")
    _ok(f"Entraînement: {temps_train_total:.1f}s ({100*temps_train_total/max(total_mesure,1):.0f}%)")
    _ok(f"Durée totale: {t_total:.1f}s  ({t_total/60:.1f} min)")


# ---------------------------------------------------------------------------
# PHASE 3 — Évaluation (Phase 10 : inclut semi-pros)
# ---------------------------------------------------------------------------

def evaluer_agent(nb_mains: int = 1000) -> dict:
    """
    Évalue AXIOM contre les 3 baselines et les 3 bots semi-pro.
    Retourne un dict avec toutes les métriques pour le journal CSV.
    """
    _section("Évaluation — Benchmark AXIOM (Phase 10)")

    from ai.agent import creer_agent
    from training.evaluator import evaluer_agent as _evaluer

    _info(f"Mains par test : {nb_mains:,}")
    _info(f"Scénarios      : Aléatoire, Call-Only, Raise-Only, TAG, LAG, Régulier")
    print()

    agent   = creer_agent(verbose=True)
    debut   = time.time()
    rapport = _evaluer(agent, nb_mains=nb_mains, verbose=True)
    duree   = time.time() - debut

    _ok(f"Évaluation terminée en {duree:.1f}s")

    return {
        # Baselines
        "winrate_vs_aleatoire"  : rapport.winrate_vs_aleatoire,
        "winrate_vs_call"       : rapport.winrate_vs_call,
        "winrate_vs_raise"      : rapport.winrate_vs_raise,
        # Semi-pros (Phase 10)
        "winrate_vs_tag"        : rapport.winrate_vs_tag,
        "winrate_vs_lag"        : rapport.winrate_vs_lag,
        "winrate_vs_regulier"   : rapport.winrate_vs_regulier,
        # Agrégats
        "winrate_moyen"         : rapport.winrate_moyen,
        "winrate_moyen_semipro" : rapport.winrate_moyen_semipro,
        "exploitabilite"        : rapport.exploitabilite_approx,
    }


def evaluer_depuis_checkpoint(nb_mains: int = 1000) -> None:
    """
    Charge les .pt du dernier checkpoint, évalue l'agent,
    et ajoute les résultats dans un fichier Excel cumulatif.
    """
    try:
        import openpyxl
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        _warn("openpyxl non installé — pip install openpyxl")
        return

    num_it, chemin_checkpoint = _trouver_dernier_checkpoint()
    if not chemin_checkpoint:
        _warn("Aucun checkpoint trouvé dans data/checkpoints/")
        return

    _section(f"Évaluation — Checkpoint iteration_{num_it:03d}")
    _info(f"Chargement des .pt depuis : {chemin_checkpoint}")

    # Copier les .pt du checkpoint dans data/models
    nb_copies = 0
    for f in os.listdir(chemin_checkpoint):
        if f.endswith(".pt"):
            shutil.copy2(os.path.join(chemin_checkpoint, f),
                         os.path.join("data", "models", f))
            nb_copies += 1
    _info(f"{nb_copies} réseaux copiés → data/models/")

    # Lancer l'évaluation complète
    resultats = evaluer_agent(nb_mains=nb_mains)

    # ── Export Excel ────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(_CHEMIN_EXCEL_EVAL), exist_ok=True)

    entetes = [
        "Itération", "Date",
        "vs Aléatoire", "vs Call-Only", "vs Raise-Only",
        "vs TAG", "vs LAG", "vs Régulier",
        "Winrate moyen", "Winrate semi-pro",
    ]

    if os.path.exists(_CHEMIN_EXCEL_EVAL):
        wb = load_workbook(_CHEMIN_EXCEL_EVAL)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Évaluations Deep CFR"

        # En-têtes stylisés
        for col, titre in enumerate(entetes, start=1):
            cell = ws.cell(row=1, column=col, value=titre)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="2F4F7F")
            cell.alignment = Alignment(horizontal="center")

        # Largeurs de colonnes
        largeurs = [12, 20, 14, 14, 14, 10, 10, 14, 16, 18]
        for col, larg in enumerate(largeurs, start=1):
            ws.column_dimensions[get_column_letter(col)].width = larg

    # Nouvelle ligne de données
    ligne = [
        num_it,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        resultats.get("winrate_vs_aleatoire", ""),
        resultats.get("winrate_vs_call",      ""),
        resultats.get("winrate_vs_raise",     ""),
        resultats.get("winrate_vs_tag",       ""),
        resultats.get("winrate_vs_lag",       ""),
        resultats.get("winrate_vs_regulier",  ""),
        resultats.get("winrate_moyen",        ""),
        resultats.get("winrate_moyen_semipro",""),
    ]

    # Colorier la ligne selon le winrate semi-pro
    winrate_sp = resultats.get("winrate_moyen_semipro", 0) or 0
    if   winrate_sp >= 10:  couleur = "C6EFCE"   # vert
    elif winrate_sp >= 0:   couleur = "FFEB9C"   # jaune
    else:                   couleur = "FFC7CE"   # rouge

    row_num = ws.max_row + 1
    for col, val in enumerate(ligne, start=1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill      = PatternFill("solid", fgColor=couleur)
        cell.alignment = Alignment(horizontal="center")

    wb.save(_CHEMIN_EXCEL_EVAL)
    _ok(f"Excel mis à jour : {_CHEMIN_EXCEL_EVAL}")
    _ok(f"Itération {num_it} | Winrate semi-pro : {winrate_sp:+.1f} bb/100")


def benchmark_rapide_agent() -> None:
    _section("Benchmark Rapide semi-pro (100 mains)")
    from ai.agent import creer_agent
    from training.evaluator import benchmark_rapide
    agent = creer_agent(verbose=False)
    wr = benchmark_rapide(agent, nb_mains=100)
    print(f"\n  Winrate semi-pro moyen : {wr:+.1f} bb/100\n")


# ---------------------------------------------------------------------------
# JOURNAL CSV
# ---------------------------------------------------------------------------

def journaliser(chemin_log: str, donnees: dict) -> None:
    os.makedirs(os.path.dirname(chemin_log), exist_ok=True)
    ecrire_entete = not os.path.exists(chemin_log)
    with open(chemin_log, 'a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=list(donnees.keys()))
        if ecrire_entete:
            writer.writeheader()
        writer.writerow(donnees)
    _info(f"Journal mis à jour : {chemin_log}")


# ---------------------------------------------------------------------------
# POINT D'ENTRÉE PRINCIPAL
# ---------------------------------------------------------------------------

def main() -> None:

    parser = argparse.ArgumentParser(
        description="AXIOM — Script d'entraînement",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--mode",
        choices=["complet", "mccfr", "deepcfr", "eval", "bench", "eval-checkpoint"],
        default="complet",
    )
    parser.add_argument("--iterations",      type=int, default=None)
    parser.add_argument("--traversees",      type=int, default=None)
    parser.add_argument("--deep-iterations", type=int, default=None,
                        dest="deep_iterations")
    parser.add_argument("--nb-mains-eval",   type=int, default=1000,
                        dest="nb_mains_eval")
    parser.add_argument(
        "--workers", type=int, default=None,
        help="Nombre de cœurs CPU (défaut : tous - 1)"
    )
    parser.add_argument(
        "--biais",
        choices=["baseline", "fold", "call", "raise"],
        default="baseline",
        help=("Point 2 — Continuation Strategies (Pluribus). Entraîne une "
              "variante biaisée du blueprint qui sera sauvegardée sous "
              "blueprint_v1_<biais>.pkl. baseline = blueprint standard.")
    )
    parser.add_argument(
        "--biais-alpha", type=float, default=0.05, dest="biais_alpha",
        help="Magnitude du biais (défaut 0.05, recommandé Pluribus)."
    )
    args = parser.parse_args()

    # Continuation Strategies : construire la config de biais propagée aux workers
    biais_conf = None
    if args.biais != "baseline":
        biais_conf = (args.biais, args.biais_alpha)

    from config.settings import (
        MCCFR_ITERATIONS, MCCFR_SAVE_EVERY,
        DEEP_CFR_ITERATIONS, DEEP_CFR_TRAVERSALS,
        CHEMIN_LOG,
    )

    iterations_mccfr = args.iterations      or MCCFR_ITERATIONS
    iterations_deep  = args.deep_iterations or DEEP_CFR_ITERATIONS
    traversees_deep  = args.traversees      or DEEP_CFR_TRAVERSALS
    nb_workers       = args.workers

    cpu_dispo         = mp.cpu_count()
    workers_effectifs = nb_workers if nb_workers else max(1, cpu_dispo - 1)

    _titre(f"AXIOM — Entraînement  [{args.mode.upper()}]")
    print(f"  Date    : {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  Mode    : {args.mode}")
    if args.mode in ("complet", "mccfr"):
        print(f"  MCCFR   : {iterations_mccfr:,} itérations sur 8 niveaux")
        print(f"  CPU     : {workers_effectifs} workers / {cpu_dispo} cœurs")
    if args.mode in ("complet", "deepcfr"):
        print(f"  DeepCFR : {iterations_deep} × {traversees_deep} traversées (GPU)")

    creer_dossiers()

    debut_global = time.time()

    # Structure CSV Phase 10 : inclut les métriques semi-pro
    metriques = {
        "date"                   : datetime.now().isoformat(timespec='seconds'),
        "mode"                   : args.mode,
        "iterations_mccfr"       : iterations_mccfr,
        "iterations_deepcfr"     : iterations_deep,
        "nb_workers"             : workers_effectifs,
        # Baselines
        "winrate_vs_aleatoire"   : "",
        "winrate_vs_call"        : "",
        "winrate_vs_raise"       : "",
        # Semi-pros
        "winrate_vs_tag"         : "",
        "winrate_vs_lag"         : "",
        "winrate_vs_regulier"    : "",
        # Agrégats
        "winrate_moyen"          : "",
        "winrate_moyen_semipro"  : "",
        "exploitabilite"         : "",
    }

    try:
        if args.mode == "eval-checkpoint":
            evaluer_depuis_checkpoint(nb_mains=args.nb_mains_eval)

        elif args.mode == "bench":
            benchmark_rapide_agent()

        elif args.mode == "eval":
            resultats = evaluer_agent(nb_mains=args.nb_mains_eval)
            metriques.update(resultats)
            journaliser(CHEMIN_LOG, metriques)

        elif args.mode == "mccfr":
            entrainer_mccfr(iterations_mccfr, MCCFR_SAVE_EVERY, nb_workers,
                            biais_conf=biais_conf)
            journaliser(CHEMIN_LOG, metriques)

        elif args.mode == "deepcfr":
            entrainer_deep_cfr(iterations_deep, traversees_deep)
            journaliser(CHEMIN_LOG, metriques)

        elif args.mode == "complet":
            entrainer_mccfr(iterations_mccfr, MCCFR_SAVE_EVERY, nb_workers,
                            biais_conf=biais_conf)
            entrainer_deep_cfr(iterations_deep, traversees_deep)
            resultats = evaluer_agent(nb_mains=args.nb_mains_eval)
            metriques.update(resultats)
            journaliser(CHEMIN_LOG, metriques)

    except KeyboardInterrupt:
        print("\n\n  Entraînement interrompu (Ctrl+C).")
        print("  Les fichiers partiels ont été sauvegardés.")

    except Exception as e:
        _warn(f"Erreur : {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    duree_globale = time.time() - debut_global
    _titre("Entraînement terminé")
    _ok(f"Durée totale : {duree_globale:.1f}s  ({duree_globale/60:.1f} min)")
    print()
    print("  Pour jouer avec la stratégie entraînée :")
    print("      python main.py")
    print()


# ---------------------------------------------------------------------------
# ENTRÉE DU SCRIPT
# IMPORTANT : if __name__ == '__main__' obligatoire sur Windows
# pour le multiprocessing (méthode spawn par défaut).
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    dossier_script = os.path.dirname(os.path.abspath(__file__))
    if dossier_script not in sys.path:
        sys.path.insert(0, dossier_script)

    main()
